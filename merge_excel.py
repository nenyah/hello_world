import openpyxl
import os

'''
@author:steven
date: 2016-05-18
'''

def get_filenames(path):
	filenames = []
	for i in os.walk(path):
		for filename in i[-1]:
			full_filename = os.path.join(i[0],filename)
			filenames.append(full_filename)
	return filenames

def read_excel(path):

	wb = openpyxl.load_workbook(path)
	ws = wb.active

	ncols = ws.max_column
	nrows= ws.max_row
	titles = []
	info = {}
	data = []
	for i in range(1, ncols+1):
		titles.append(ws.cell(row=1,column=i).value)
	# print(titles)

	for row in range(2,nrows+1):
		temp_list = []
		for col in range(1, ncols+1):
			temp_list.append(ws.cell(row=row,column=col).value)
		data.append(temp_list)
	# print(data)
	return data,titles


def save_excel(data,titles,path = None):
	if path is None:
		path = 'Total.xlsx'

	wb = openpyxl.Workbook()
	ws = wb.active

	for index, title in enumerate(titles):
		ws.cell(row=1,column=index+1,value=title)
	for row,item in enumerate(data):
		for col,value in enumerate(item):
			ws.cell(row=row+2,column=col+1,value=value)

	wb.save(path)

if __name__ == '__main__':
	print('Program is running...')
	path = r'E:\Work\06-Work\00-Todo\Self Calc\Freight Self Calc Download'
	target_path = r'E:\Work\06-Work\00-Todo\Self Calc\Freight Self Calc Sort'
	data = []
	titles = ''
	for filename in get_filenames(path):
		titles = read_excel(filename)[-1]
		for item in read_excel(filename)[0]:
			if item not in data:
				data.append(item)
	save_excel(data,titles,target_path+os.sep+'Total.xlsx')
	print('Success!')