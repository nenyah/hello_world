import openpyxl
import xlrd
import os
'''
@author:steven
date: 2016-05-16
'''
def get_ref():
	"""read the ref data
	Keyword arguments: 
	----------------------
	path default(freight_standard.xls)
	Returns: DataFrame
	"""
	path = r'freight_standard.xls'
	wb = xlrd.open_workbook(path)
	sheet = wb.sheet_by_name('register')
	nrows = sheet.nrows
	ref = []
	for i in range(1, nrows):
		if sheet.cell_value(i,2) != 42:
			data = {
			'country': sheet.cell_value(i,2),
			'price': sheet.cell_value(i,4)
			}
			ref.append(data)
	return ref
# 一次性加入内存，减少读入次数
ref = get_ref()

# 按规则计算得到运算结果
def calc(country,weight,freight):
	"""calc the price form parameters
	Parameters
	----------
	country: str country value
	weight: float weight value
	freight: str freight value

	Returns
	-------
	float the price of the calc result
	"""
	result = 0
	if '小包' in freight:
		freight = 'cn_post'
	elif '专线' in freight:
		freight = 'ru_express'
	elif '宝' in freight:
		freight = 'eub'
	elif '苍南' or '昆山' in freight:
		freight = 'electric'
	# print(freight)
	discount = {
	'cn_post': 0.82,
	'ru_express': 0.72,
	'electric': 0.77
	}

	if freight in ['cn_post','ru_express','electric']:
		for item in ref:
			if country == item['country']:
				try:
					result = (item['price']*10**-3*weight+8)*discount[freight]
				except TypeError:
					print('CN_POST weight is empty')
	elif freight == 'eub':
		if country == 'United States':
			try:
				if weight < 70:
					result = 70 * 0.08 + 9
				elif weight <= 200:
					result = weight * 0.08 + 9
				elif weight < 2000:
					result = weight * 0.075 + 9
				# print(weight,result)
			except TypeError:
				print('EUB weight is empty')
		elif country == 'Russian Federation':
			result = weight * 0.08 + 8
		else:
			result = weight*0.07 + 22
	else:
		pass
	return round(result,2)


def save_excel(source_path,target_path=None):
	"""compute data from source file and generate a new file
	Parameters
	----------
	source_path: file
	target_path: file
	"""
	if target_path == None:
		target_path = os.path.splitext(source_path)[0]+'_res.xlsx'
	wb = openpyxl.load_workbook(source_path)
	ws = wb.active

	ncols = ws.max_column
	nrows= ws.max_row

	channel_col = 0
	track_col = 0
	country_col = 0
	weight_col = 0
	for i in range(1,ncols+1):
		value = ws.cell(row=1,column=i).value
		if '渠道' in value:
			channel_col = i
		elif '单号' in value:
			track_col = i
		elif '国家' in value:
			country_col = i
		elif '重量' in value:
			weight_col = i

	save_book = openpyxl.Workbook()
	save_sheet = save_book.active
	save_sheet.title = 'freight calc'

	titles = ['渠道','国家','物流单号','结算重量','结算运费']
	for index,item in enumerate(titles):
		save_sheet.cell(row=1,column=index+1,value=item)
	counter = 2
	for row in range(2,nrows+1):
		channel = ws.cell(row=row,column=channel_col).value
		track = ws.cell(row=row,column=track_col).value
		country = ws.cell(row=row,column=country_col).value
		weight = ws.cell(row=row,column=weight_col).value
		result = calc(country,weight,channel)
		data = [channel,country,track,weight,result]
		for index, item in enumerate(data):
			save_sheet.cell(row=counter,column=index+1,value=item)
		counter += 1
		print(channel,track,country,weight,result)

	save_book.save(target_path)

if __name__ == '__main__':
	path = r'E:\Work\06-Work\00-Todo\Self Calc\Freight Self Calc Sort\Total.xlsx'
	save_excel(path)